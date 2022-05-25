
# this script: 
#
# gets search parameters from one excel file 
# makes an url for the google search pages using the words in the excel file
#       search for example "personname" AND "word1" OR "word2" OR ... OR "word-n"
# opens google search pages on webbrowser tabs and takes screenshots of them.
#         -screenshots will be taken of the whole screen, so open a webbrowser window to full the screen
# 
# saves results as pictures in a subfolder
# saves 1 excel for each person name telling what search was done and when

# this script has almost no error handling (yet). 
# It requires an input at the end, so you know that if you get the line asking for the end input, 
# the program didn't crash. If you don't get it, something went wrong.


# importing the modules
import requests
import os
import webbrowser
import openpyxl
from datetime import date, datetime
import time
from openpyxl.utils import get_column_letter
import pyautogui
import PIL


#print telling the user that the program is running.
print("Program is running... Please wait \n ... \n ...")

###########################################
# open the excel with the search key words
###########################################

# Give the location of the file.
path_words = "search-keywords.xlsx"
# To open the workbook workbook object is created
wb_obj_words = openpyxl.load_workbook(path_words)
# Get workbook active sheet object from the active attribute
sheet_obj_words = wb_obj_words.active
#Determine total number of rows 
number_of_word_file_rows = sheet_obj_words.max_row

###########################################
# get the words for the searches
###########################################

#get person list (max 10)
key_persons_list = []
the_row = 4
for i in range(0,10):
    cell_obj_pers = sheet_obj_words.cell(row = the_row, column = 2)
    pers = cell_obj_pers.value
    key_person = str(pers)
    if key_person != "None":
        key_persons_list.append(key_person)
        the_row += 1
    else:
        break

#get language selection
cell_obj_lan = sheet_obj_words.cell(row = 4, column = 3)
lan = cell_obj_lan.value
language = str(lan)

#get key words list (max 11), with the chosen language
keyword_list_with_chosen_language = []
the_row = 4

#english US
if language == "en":
    for i in range(0,11):
        cell_obj_word = sheet_obj_words.cell(row = the_row, column = 4)
        word = cell_obj_word.value
        keyword = str(word)
        if keyword != "None":
            keyword_list_with_chosen_language.append(keyword)
            the_row += 1
        else:
            break

#lithuanian
if language == "fi":
    for i in range(0,11):
        cell_obj_word = sheet_obj_words.cell(row = the_row, column = 5)
        word = cell_obj_word.value
        keyword = str(word)
        if keyword != "None":
            keyword_list_with_chosen_language.append(keyword)
            the_row += 1
        else:
            break

##############################################
#make a subfolder for the results
##############################################

#folder name: "results" and date

datenow = str(date.today())
datetimenow = datetime.now()
timenow = datetimenow.strftime("%H-%M-%S")

folder_name = f'results-{datenow}'
try:
    os.mkdir(folder_name)
except FileExistsError:
    reply = str(input("folder already exists. Do you want to write over the files in the current folder? y/n "))
    if reply.lower() == "y":
        pass
    else:
        folder_name += f"-{str(timenow)}"
        os.mkdir(folder_name)

################################################################################################################
# make new excel files for the key person results
################################################################################################################

#names of the files: person-name_language_googlesearch_date-now

#FOR EACH KEY PERSON:
for x in range(0, len(key_persons_list)):
    #make the result file for person number x
    person_no_space = str(key_persons_list[x]).replace(" ", "-")
    name_for_keyperson_result_file = f"{person_no_space}_{language}_googlesearch_{datenow}.xlsx"
    wb_keyperson = openpyxl.Workbook()
    wb_keyperson.save(f'.//{folder_name}//{name_for_keyperson_result_file}')
    sheet_obj_keyperson_result = wb_keyperson.active

    ######################################################
    # add the headers to keyperson[x] result file
    ######################################################
    c1 = sheet_obj_keyperson_result.cell(row = 1, column = 1)
    c1.value = "Person name"
    c2 = sheet_obj_keyperson_result.cell(row = 2, column = 1)
    c2.value = "Language of the search"
    c3 = sheet_obj_keyperson_result.cell(row = 3, column = 1)
    c3.value = "keywords"
    c4 = sheet_obj_keyperson_result.cell(row = 4, column = 1)
    c4.value = "Date of the search"
    c5 = sheet_obj_keyperson_result.cell(row = 6, column = 1)
    c5.value = "Google search page URL"
    c6 = sheet_obj_keyperson_result.cell(row = 8, column = 1)
    c6.value = "Google search results saved as screenshots of the search page"
    #save the file
    wb_keyperson.save(f'.//{folder_name}//{name_for_keyperson_result_file}')

    ##############################################################################
    # add search info to result file (company name, language, keywords, date)
    ##############################################################################
    c1 = sheet_obj_keyperson_result.cell(row = 1, column = 2)
    c1.value = str(key_persons_list[x])
    c2 = sheet_obj_keyperson_result.cell(row = 2, column = 2)
    c2.value = str(language)
    c3 = sheet_obj_keyperson_result.cell(row = 3, column = 2)
    c3.value = str(keyword_list_with_chosen_language)
    c4 = sheet_obj_keyperson_result.cell(row = 4, column = 2)
    c4.value = str(datenow)
    #save the file
    wb_keyperson.save(f'.//{folder_name}//{name_for_keyperson_result_file}')

    ###############################################################################################################
    # do the keyperson[x] search, open the search result page and add the URL results to keyperson[x] result file
    ###############################################################################################################

    url_list = []
    # stored queries in a list
    query_list = keyword_list_with_chosen_language
    # save the company name in a variable and  replace spaces with %20
    keyperson_name_no_spaces = str(key_persons_list[x]).replace(" ", "%20")

    number_of_keywords = len(query_list)
    the_query = f'"{key_persons_list[x]}" AND '
    search_page_url = 'https://google.com/search?q=' + f'%22{keyperson_name_no_spaces}%22%20AND%20'
    # iterate through different keywords, add them to the query & search url
    for j in query_list:
        the_query += f'"{j}"'
        j_no_spaces = str(j).replace(" ", "%20") #replace spaces with %20
        search_page_url += f'%22{j_no_spaces}%22'
        if j != query_list[-1]:
            the_query += ' OR '
            search_page_url += '%20OR%20'
        else:
            pass
    
    search_page_url += f'&lr=lang_{language}'
    #open the search result page in a default browser tab
    try:
        time.sleep(5)
        webbrowser.open(search_page_url, new=2)
    except:
        print("was not able to open the web page for some reason")

    #wait some time for the webpage to open
    time.sleep(5)

    #save the screenshot to the subfolder
    myscreenshot3 = pyautogui.screenshot(f'.//{folder_name}//screenshot-{person_no_space}.png')

    #add the search page url to the company result file
    c2 = sheet_obj_keyperson_result.cell(row = 6, column = 2)
    c2.value = str(search_page_url)

    ##########################
    
    #adjust the width of the column
    sheet_obj_keyperson_result.column_dimensions[get_column_letter(1)].width = 40

    wb_keyperson.save(f'.//{folder_name}//{name_for_keyperson_result_file}')

#open a webpage to tell user the program is finished, screenshots are taken and it's safe to touch the browser
webbrowser.open('https://pexels.com/search/cat/', new=2)

#  a print telling the user the program is finished
#  response takes enter as an input to end the program
print("Program is finished, results have been saved")
response = input("Press enter to close this window and end the program. \n ...")
